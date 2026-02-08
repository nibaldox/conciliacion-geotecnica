"""Export comparison results to formatted Excel workbook."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Style constants
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496",
                          fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center",
                         wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
FILL_OK = PatternFill(start_color="C6EFCE", end_color="C6EFCE",
                      fill_type="solid")
FILL_WARN = PatternFill(start_color="FFEB9C", end_color="FFEB9C",
                        fill_type="solid")
FILL_NOK = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                       fill_type="solid")
FONT_OK = Font(color="006100")
FONT_WARN = Font(color="9C5700")
FONT_NOK = Font(color="9C0006")


def _apply_status_style(cell):
    """Apply conditional formatting based on status text."""
    val = cell.value
    if val == "CUMPLE":
        cell.fill = FILL_OK
        cell.font = FONT_OK
    elif val == "FUERA DE TOLERANCIA":
        cell.fill = FILL_WARN
        cell.font = FONT_WARN
    elif val == "NO CUMPLE":
        cell.fill = FILL_NOK
        cell.font = FONT_NOK


def _write_header(ws, row, headers):
    """Write a styled header row."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def _auto_width(ws):
    """Auto-adjust column widths."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 25)


def _write_summary_sheet(wb, comparisons, tolerances, project_info):
    """Create the Resumen sheet."""
    ws = wb.active
    ws.title = "Resumen"

    # Project info
    row = 1
    ws.cell(row=row, column=1,
            value="CONCILIACION GEOTECNICA: DISENO vs AS-BUILT").font = Font(
        bold=True, size=14, color="2F5496")
    row = 3
    info_fields = [
        ("Proyecto", project_info.get('project', '')),
        ("Operacion", project_info.get('operation', '')),
        ("Fase / Pit", project_info.get('phase', '')),
        ("Elaborado por", project_info.get('author', '')),
        ("Fecha", project_info.get('date', '')),
    ]
    for label, value in info_fields:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1

    # Tolerance table
    ws.cell(row=row, column=1,
            value="Tolerancias Aplicadas").font = Font(bold=True, size=12)
    row += 1
    _write_header(ws, row, ["Parametro", "Tol. (-)", "Tol. (+)"])
    row += 1
    tol_display = [
        ("Altura de banco (m)", 'bench_height'),
        ("Angulo cara (deg)", 'face_angle'),
        ("Ancho de berma (m)", 'berm_width'),
        ("Angulo inter-rampa (deg)", 'inter_ramp_angle'),
        ("Angulo global (deg)", 'overall_angle'),
    ]
    for label, key in tol_display:
        tol = tolerances.get(key, {'neg': 0, 'pos': 0})
        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws.cell(row=row, column=2, value=tol['neg']).border = THIN_BORDER
        ws.cell(row=row, column=3, value=tol['pos']).border = THIN_BORDER
        row += 1

    row += 1

    # Compliance summary
    if comparisons:
        ws.cell(row=row, column=1,
                value="Resumen de Cumplimiento").font = Font(
            bold=True, size=12)
        row += 1
        _write_header(ws, row,
                      ["Parametro", "CUMPLE", "FUERA TOL.", "NO CUMPLE",
                       "Total", "% Cumpl."])
        row += 1

        for key, label in [('height_status', 'Altura de banco'),
                           ('angle_status', 'Angulo de cara'),
                           ('berm_status', 'Ancho de berma')]:
            total = len(comparisons)
            n_ok = sum(1 for c in comparisons if c[key] == "CUMPLE")
            n_warn = sum(1 for c in comparisons
                        if c[key] == "FUERA DE TOLERANCIA")
            n_nok = sum(1 for c in comparisons if c[key] == "NO CUMPLE")
            pct = n_ok / total * 100 if total > 0 else 0

            ws.cell(row=row, column=1, value=label).border = THIN_BORDER
            c_ok = ws.cell(row=row, column=2, value=n_ok)
            c_ok.border = THIN_BORDER
            c_ok.fill = FILL_OK
            c_warn = ws.cell(row=row, column=3, value=n_warn)
            c_warn.border = THIN_BORDER
            c_warn.fill = FILL_WARN
            c_nok = ws.cell(row=row, column=4, value=n_nok)
            c_nok.border = THIN_BORDER
            c_nok.fill = FILL_NOK
            ws.cell(row=row, column=5, value=total).border = THIN_BORDER
            ws.cell(row=row, column=6,
                    value=f"{pct:.1f}%").border = THIN_BORDER
            row += 1

    _auto_width(ws)


def _write_bench_sheet(wb, comparisons):
    """Create the Bancos detail sheet."""
    ws = wb.create_sheet("Bancos")

    headers = [
        "Sector", "Seccion", "Banco", "Nivel",
        "H. Diseno (m)", "H. Real (m)", "Desv. H (m)", "Cumpl. H",
        "A. Diseno (deg)", "A. Real (deg)", "Desv. A (deg)", "Cumpl. A",
        "B. Diseno (m)", "B. Real (m)", "Desv. B (m)", "Cumpl. B",
    ]
    _write_header(ws, 1, headers)

    for row_idx, comp in enumerate(comparisons, 2):
        values = [
            comp['sector'], comp['section'], comp['bench_num'], comp['level'],
            comp['height_design'], comp['height_real'],
            comp['height_dev'], comp['height_status'],
            comp['angle_design'], comp['angle_real'],
            comp['angle_dev'], comp['angle_status'],
            comp['berm_design'], comp['berm_real'],
            comp['berm_dev'], comp['berm_status'],
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            if col_idx in (8, 12, 16):
                _apply_status_style(cell)

    _auto_width(ws)


def _write_interramp_sheet(wb, params_design, params_topo):
    """Create the Inter-Rampa sheet."""
    ws = wb.create_sheet("Inter-Rampa")

    headers = [
        "Seccion", "Sector",
        "Ang. IR Diseno (deg)", "Ang. IR Real (deg)",
        "Ang. Global Diseno (deg)", "Ang. Global Real (deg)",
    ]
    _write_header(ws, 1, headers)

    n = min(len(params_design), len(params_topo))
    for i in range(n):
        pd = params_design[i]
        pt = params_topo[i]
        row = i + 2
        ws.cell(row=row, column=1, value=pd.section_name).border = THIN_BORDER
        ws.cell(row=row, column=2, value=pd.sector).border = THIN_BORDER
        ws.cell(row=row, column=3,
                value=round(pd.inter_ramp_angle, 1)).border = THIN_BORDER
        ws.cell(row=row, column=4,
                value=round(pt.inter_ramp_angle, 1)).border = THIN_BORDER
        ws.cell(row=row, column=5,
                value=round(pd.overall_angle, 1)).border = THIN_BORDER
        ws.cell(row=row, column=6,
                value=round(pt.overall_angle, 1)).border = THIN_BORDER

    _auto_width(ws)


def _write_dashboard_sheet(wb, comparisons):
    """Create a simple Dashboard summary sheet."""
    ws = wb.create_sheet("Dashboard")

    ws.cell(row=1, column=1,
            value="DASHBOARD DE CUMPLIMIENTO").font = Font(
        bold=True, size=14, color="2F5496")

    if not comparisons:
        ws.cell(row=3, column=1, value="Sin datos de comparacion.")
        return

    total = len(comparisons)
    row = 3
    _write_header(ws, row,
                  ["Parametro", "CUMPLE", "FUERA TOL.", "NO CUMPLE",
                   "% Cumplimiento"])
    row += 1

    for key, label in [('height_status', 'Altura de banco'),
                       ('angle_status', 'Angulo de cara'),
                       ('berm_status', 'Ancho de berma')]:
        n_ok = sum(1 for c in comparisons if c[key] == "CUMPLE")
        n_warn = sum(1 for c in comparisons
                    if c[key] == "FUERA DE TOLERANCIA")
        n_nok = sum(1 for c in comparisons if c[key] == "NO CUMPLE")
        pct = n_ok / total * 100 if total > 0 else 0

        ws.cell(row=row, column=1, value=label).border = THIN_BORDER
        c_ok = ws.cell(row=row, column=2, value=n_ok)
        c_ok.border = THIN_BORDER
        c_ok.fill = FILL_OK
        c_warn = ws.cell(row=row, column=3, value=n_warn)
        c_warn.border = THIN_BORDER
        c_warn.fill = FILL_WARN
        c_nok = ws.cell(row=row, column=4, value=n_nok)
        c_nok.border = THIN_BORDER
        c_nok.fill = FILL_NOK
        ws.cell(row=row, column=5,
                value=f"{pct:.1f}%").border = THIN_BORDER
        row += 1

    # Global compliance
    row += 1
    n_all = total * 3
    n_ok_all = sum(
        1 for c in comparisons
        for k in ['height_status', 'angle_status', 'berm_status']
        if c[k] == "CUMPLE"
    )
    pct_global = n_ok_all / n_all * 100 if n_all > 0 else 0

    ws.cell(row=row, column=1,
            value="CUMPLIMIENTO GLOBAL").font = Font(bold=True, size=12)
    ws.cell(row=row, column=2,
            value=f"{pct_global:.1f}%").font = Font(bold=True, size=14)

    _auto_width(ws)


def export_results(comparisons, params_design, params_topo,
                   tolerances, output_path, project_info=None):
    """Export comparison results to a formatted Excel workbook."""
    if project_info is None:
        project_info = {}

    wb = openpyxl.Workbook()

    _write_summary_sheet(wb, comparisons, tolerances, project_info)
    _write_bench_sheet(wb, comparisons)
    _write_interramp_sheet(wb, params_design, params_topo)
    _write_dashboard_sheet(wb, comparisons)

    wb.save(output_path)
